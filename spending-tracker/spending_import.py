import csv
import hashlib
import io
import json
import re
import unicodedata
from datetime import datetime

from models import (db, Category, Transaction, Import, CategorizationRule,
                    get_uncategorized_category, get_default_income_category, get_category)

MAX_FILE_BYTES = 2 * 1024 * 1024  # 2MB
MAX_ROWS = 10000

# Hebrew exports embed bidi control characters inside date and amount cells.
# They are NOT whitespace, so .strip() leaves them in place and both strptime()
# and float() then fail on what looks like a perfectly ordinary value.
_CELL_TRANSLATE = {c: None for c in (
    0x200E, 0x200F,                             # LRM, RLM
    0x202A, 0x202B, 0x202C, 0x202D, 0x202E,     # embeddings / overrides
    0x2066, 0x2067, 0x2068, 0x2069,             # isolates
    0xFEFF,                                     # BOM / zero-width no-break space
)}
_CELL_TRANSLATE[0x00A0] = 0x20      # non-breaking space -> ordinary space
# Israeli statements mix gershayim, a modifier double-acute and ASCII quotes in
# the same document, so normalise them before comparing marker text.
_CELL_TRANSLATE.update({0x05F4: 0x22, 0x02DD: 0x22, 0x201C: 0x22, 0x201D: 0x22,
                        0x05F3: 0x27, 0x2018: 0x27, 0x2019: 0x27})


def _clean(value):
    """Normalise a raw cell into comparable, parseable text."""
    if value is None:
        return ''
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return unicodedata.normalize('NFKC', str(value)).translate(_CELL_TRANSLATE).strip()


# Two-digit-year formats come last so a four-digit year always wins the match.
# Isracard emits dd.mm.yy, which no four-digit pattern can parse.
DATE_FORMATS = [
    '%Y-%m-%d', '%d/%m/%Y', '%m/%d/%Y', '%d-%m-%Y', '%m-%d-%Y', '%Y/%m/%d', '%d.%m.%Y',
    '%Y-%m-%d %H:%M:%S', '%d/%m/%Y %H:%M',
    '%d/%m/%y', '%d.%m.%y', '%d-%m-%y',
]

_PROCESSOR_PREFIX = re.compile(r'^(SQ|TST|PAYPAL|PY|SP|IN)\s*\*\s*', re.IGNORECASE)
_CITY_STATE_SUFFIX = re.compile(r'\s+(?:[A-Z]+\s+){1,2}[A-Z]{2}$')  # e.g. " SEATTLE WA"
_STORE_NUMBER = re.compile(r'#?\d{3,}\b')
_EMBEDDED_DATE = re.compile(r'\b\d{2}/\d{2}\b')

# Branch/location suffixes worth dropping so the same chain groups as one merchant
# ("Shufersal Deal Tel Aviv" and "Shufersal Deal Haifa" are both just Shufersal Deal).
_LOCALITIES = {
    'TEL AVIV', 'JERUSALEM', 'HAIFA', 'RISHON LEZION', 'RISHON LEZIYON', 'PETAH TIKVA',
    'ASHDOD', 'NETANYA', 'BEER SHEVA', 'BEERSHEVA', 'HOLON', 'BNEI BRAK', 'RAMAT GAN',
    'RAMAT HASHARON', 'HERZLIYA', 'KFAR SABA', 'RAANANA', 'MODIIN', 'ASHKELON',
    'BAT YAM', 'REHOVOT', 'GIVATAYIM', 'HOD HASHARON', 'EILAT', 'NAHARIYA', 'HADERA',
    'LOD', 'RAMLA', 'ROSH HAAYIN', 'YAVNE', 'KIRYAT ONO', 'KIRYAT GAT', 'AFULA',
    'TIBERIAS', 'NAZARETH', 'ACRE', 'AKKO', 'DIMONA', 'YEHUD', 'OR YEHUDA', 'AZOR',
    'DIZENGOFF', 'AYALON', 'GRAND KANYON', 'KANYON',
    'תל אביב', 'ירושלים', 'חיפה', 'ראשון לציון', 'פתח תקווה', 'אשדוד', 'נתניה',
    'באר שבע', 'חולון', 'בני ברק', 'רמת גן', 'הרצליה', 'כפר סבא', 'רעננה', 'מודיעין',
}
_MAX_LOCALITY_WORDS = 2

# Keyword -> category seeded on first run so the very first import is mostly
# pre-sorted. User corrections create higher-priority rules that override these.
SEED_RULES = {
    # Groceries
    'SHUFERSAL': 'Groceries', 'RAMI LEVY': 'Groceries', 'VICTORY': 'Groceries',
    'YOHANANOF': 'Groceries', 'TIV TAAM': 'Groceries', 'OSHER AD': 'Groceries',
    'AM PM': 'Groceries', 'SUPERMARKET': 'Groceries', 'שופרסל': 'Groceries',
    'רמי לוי': 'Groceries', 'יוחננוף': 'Groceries', 'ויקטורי': 'Groceries',
    'WHOLE FOODS': 'Groceries', 'TRADER JOE': 'Groceries',
    # Restaurants / cafés / delivery
    'AROMA': 'Cafés', 'ARCAFFE': 'Cafés', 'LANDWER': 'Cafés', 'CAFE': 'Cafés',
    'COFFEE': 'Cafés', 'STARBUCKS': 'Cafés', 'קפה': 'Cafés',
    'RESTAURANT': 'Restaurants', 'PIZZA': 'Restaurants', 'SUSHI': 'Restaurants',
    'BURGER': 'Restaurants', 'MCDONALD': 'Restaurants', 'מסעדה': 'Restaurants',
    'WOLT': 'Food Delivery', '10BIS': 'Food Delivery', 'TENBIS': 'Food Delivery',
    'תן ביס': 'Food Delivery',
    # Car
    'PAZ': 'Fuel', 'DELEK': 'Fuel', 'SONOL': 'Fuel', 'DOR ALON': 'Fuel',
    'GAS STATION': 'Fuel', 'SHELL': 'Fuel', 'דלק': 'Fuel', 'פז': 'Fuel',
    'PANGO': 'Parking', 'CELLOPARK': 'Parking', 'PARKING': 'Parking', 'חניון': 'Parking',
    'KVISH 6': 'Tolls', 'KVISH6': 'Tolls', 'כביש 6': 'Tolls',
    'GARAGE': 'Car Maintenance', 'מוסך': 'Car Maintenance',
    # Cash
    'ATM': 'Cash Withdrawal', 'CASH WITHDRAWAL': 'Cash Withdrawal',
    'WITHDRAWAL': 'Cash Withdrawal', 'כספומט': 'Cash Withdrawal', 'משיכת מזומן': 'Cash Withdrawal',
    # Home / utilities
    'BEZEQ': 'Internet & TV', 'HOT ': 'Internet & TV', 'YES ': 'Internet & TV',
    'PARTNER': 'Internet & TV', 'CELLCOM': 'Internet & TV', 'PELEPHONE': 'Internet & TV',
    'בזק': 'Internet & TV', 'ELECTRIC': 'Utilities', 'CHEVRAT HASHMAL': 'Utilities',
    'חשמל': 'Utilities', 'MEKOROT': 'Utilities', 'מים': 'Utilities',
    'ARNONA': 'Utilities', 'ארנונה': 'Utilities',
    'RENT': 'Rent / Mortgage', 'MORTGAGE': 'Rent / Mortgage', 'MASHKANTA': 'Rent / Mortgage',
    'משכנתא': 'Rent / Mortgage', 'שכר דירה': 'Rent / Mortgage',
    'IKEA': 'Furniture & Appliances', 'ACE ': 'Home Maintenance', 'HOME CENTER': 'Home Maintenance',
    # Transport
    'RAV KAV': 'Public Transport', 'RAVKAV': 'Public Transport', 'רב קו': 'Public Transport',
    'EGGED': 'Public Transport', 'ISRAEL RAILWAYS': 'Public Transport', 'רכבת': 'Public Transport',
    'GETT': 'Taxi & Rideshare', 'UBER': 'Taxi & Rideshare', 'YANGO': 'Taxi & Rideshare',
    'MONIT': 'Taxi & Rideshare', 'מונית': 'Taxi & Rideshare',
    # Health
    'SUPER PHARM': 'Pharmacy', 'SUPERPHARM': 'Pharmacy', 'PHARMACY': 'Pharmacy',
    'סופר פארם': 'Pharmacy', 'בית מרקחת': 'Pharmacy',
    'CLALIT': 'Health Insurance', 'MACCABI': 'Health Insurance', 'MEUHEDET': 'Health Insurance',
    'כללית': 'Health Insurance', 'מכבי': 'Health Insurance',
    'HOLMES PLACE': 'Gym & Fitness', 'GYM': 'Gym & Fitness', 'חדר כושר': 'Gym & Fitness',
    # Subscriptions / entertainment
    'NETFLIX': 'Streaming & Subscriptions', 'SPOTIFY': 'Streaming & Subscriptions',
    'YOUTUBE': 'Streaming & Subscriptions', 'ICLOUD': 'Streaming & Subscriptions',
    'APPLE.COM': 'Streaming & Subscriptions', 'GOOGLE STORAGE': 'Streaming & Subscriptions',
    'DISNEY': 'Streaming & Subscriptions', 'CINEMA CITY': 'Events & Nightlife',
    'YES PLANET': 'Events & Nightlife',
    # Shopping
    'AMAZON': 'General Shopping', 'ALIEXPRESS': 'General Shopping', 'EBAY': 'General Shopping',
    'ZARA': 'Clothing', 'CASTRO': 'Clothing', 'FOX ': 'Clothing', 'H&M': 'Clothing',
    'KSP': 'Electronics', 'BUG ': 'Electronics', 'IVORY': 'Electronics',
    # Travel
    'BOOKING.COM': 'Hotels', 'AIRBNB': 'Hotels', 'EL AL': 'Flights', 'ELAL': 'Flights',
    'RYANAIR': 'Flights', 'WIZZ': 'Flights', 'ISRAIR': 'Flights',
    # Financial
    'BANK FEE': 'Bank Fees', 'AMLAT': 'Bank Fees', 'עמלה': 'Bank Fees',
    'MAS HACHNASA': 'Taxes', 'מס הכנסה': 'Taxes', 'BITUACH LEUMI': 'Taxes',
    'ביטוח לאומי': 'Taxes',
    # Income
    'SALARY': 'Salary', 'MASKORET': 'Salary', 'משכורת': 'Salary', 'PAYROLL': 'Salary',
    'BONUS': 'Bonus', 'בונוס': 'Bonus', 'DIVIDEND': 'Dividends', 'INTEREST': 'Interest',

    # Hebrew merchant names — statements from Israeli issuers list businesses in
    # Hebrew, so the Latin spellings above never match them.
    'ארומה': 'Cafés', 'לנדוור': 'Cafés', 'קפה גרג': 'Cafés', 'קפה קפה': 'Cafés',
    'מקדונלד': 'Restaurants', 'בורגר': 'Restaurants', 'פיצה': 'Restaurants',
    'סושי': 'Restaurants', 'פלאפל': 'Restaurants', 'חומוס': 'Restaurants',
    'וולט': 'Food Delivery', 'טיב טעם': 'Groceries', 'אושר עד': 'Groceries',
    'יינות ביתן': 'Groceries', 'מגה': 'Groceries', 'סופרמרקט': 'Groceries',
    'סונול': 'Fuel', 'דור אלון': 'Fuel', 'טן': 'Fuel',
    'ניו פארם': 'Pharmacy', 'איקאה': 'Furniture & Appliances', 'הום סנטר': 'Home Maintenance',
    'קסטרו': 'Clothing', 'פוקס': 'Clothing', 'גולף': 'Clothing', 'רנואר': 'Clothing',
    'זארה': 'Clothing', 'קיי אס פי': 'Electronics', 'באג': 'Electronics',
    'אגד': 'Public Transport', 'דן': 'Public Transport', 'רכבת ישראל': 'Public Transport',
    'גט': 'Taxi & Rideshare', 'יאנגו': 'Taxi & Rideshare',
    'חברת חשמל': 'Utilities', 'מי אביבים': 'Utilities', 'מקורות': 'Utilities',
    'הוט': 'Internet & TV', 'יס': 'Internet & TV', 'סלקום': 'Internet & TV',
    'פלאפון': 'Internet & TV', 'פרטנר': 'Internet & TV',
    'סינמה סיטי': 'Events & Nightlife', 'יס פלאנט': 'Events & Nightlife',
    'נטפליקס': 'Streaming & Subscriptions', 'ספוטיפיי': 'Streaming & Subscriptions',
    'חדר כושר': 'Gym & Fitness', 'הולמס פלייס': 'Gym & Fitness',
    'אל על': 'Flights', 'ישראייר': 'Flights', 'ארקיע': 'Flights',
}


class ImportError_(Exception):
    pass


def seed_default_rules(user):
    """Add any built-in keyword rules the user is missing.

    Additive rather than once-only, so a release that ships new merchants
    reaches existing users too. Patterns the user already has are left alone —
    their own corrections (priority 10) always win.
    """
    existing = {r.match_pattern for r in CategorizationRule.query.filter_by(user_id=user.id).all()}
    added = False
    for pattern, category_name in SEED_RULES.items():
        if pattern in existing:
            continue
        category = get_category(user, category_name)
        if not category:
            continue
        db.session.add(CategorizationRule(
            user_id=user.id, category_id=category.id,
            match_pattern=pattern, match_type='contains', priority=0,
        ))
        added = True
    if added:
        db.session.commit()


# Column names seen across Israeli issuers. Used both to locate the header row
# and to fall back when one workbook mixes tables with different headings —
# Isracard's foreign-currency sheet uses different names from its domestic one.
COLUMN_ALIASES = {
    'date': ('תאריך רכישה', 'תאריך עסקה', 'תאריך קניה', 'תאריך חיוב', 'תאריך',
             'date', 'transaction date', 'purchase date'),
    'amount': ('סכום חיוב', 'סכום לחיוב', 'סכום עסקה', 'סכום מקורי', 'סכום',
             'בחובה', 'בזכות', 'amount', 'debit', 'credit', 'charge', 'sum'),
    'merchant': ('שם בית העסק', 'שם בית עסק', 'בית עסק', 'שם בית העסק/תיאור',
             'פירוט נוסף', 'פרוט נוסף', 'תיאור', 'סוג תנועה',
             'description', 'merchant', 'payee', 'business', 'details'),
}
HEADER_HINTS = tuple(h for group in COLUMN_ALIASES.values() for h in group)

# Summary and pending sections that sit inside the same sheet as real rows.
SKIP_ROW_MARKERS = ('סה"כ', 'סך הכל', 'עסקאות שטרם נקלטו', 'total')

# Signatures pandas uses to identify real spreadsheets; anything else with an
# .xls name is usually HTML or SpreadsheetML wearing a costume.
_XLS_SIGNATURES = (
    b"\x09\x00\x04\x00\x07\x00\x10\x00", b"\x09\x02\x06\x00\x00\x00\x10\x00",
    b"\x09\x04\x06\x00\x00\x10\x00", b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1",
)
_ZIP_SIGNATURE = b"PK\x03\x04"


def _is_xlsx(file_bytes):
    return file_bytes[:4] == _ZIP_SIGNATURE


def _is_xls(file_bytes):
    return any(file_bytes.startswith(sig) for sig in _XLS_SIGNATURES)


def _is_markup(file_bytes):
    """Some issuers (Bank Leumi) export an HTML table named .xls."""
    head = file_bytes[:4096].lstrip(b'\xef\xbb\xbf \t\r\n').lower()
    return head.startswith(b'<') and (b'<html' in head or b'<table' in head or b'<?xml' in head)


def _cell_to_text(value):
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return _clean(value)


def _decode_hebrew(file_bytes):
    """Hebrew HTML exports are frequently CP1255 rather than UTF-8."""
    for encoding in ('utf-8-sig', 'utf-8', 'windows-1255', 'cp1255'):
        try:
            return file_bytes.decode(encoding)
        except (UnicodeDecodeError, LookupError):
            continue
    return file_bytes.decode('utf-8', errors='replace')


def _read_markup_rows(file_bytes):
    """Extract the largest <table> from an HTML-disguised export."""
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        raise ImportError_('This file is an HTML table saved as .xls, and the '
                           'server is missing beautifulsoup4 to read it')
    soup = BeautifulSoup(_decode_hebrew(file_bytes), 'html.parser')
    tables = soup.find_all('table')
    if not tables:
        raise ImportError_('No table found inside the file')
    # Leumi marks the real one with class="xlTable"; otherwise take the biggest.
    table = (soup.find('table', class_='xlTable')
             or max(tables, key=lambda t: len(t.find_all('tr'))))
    rows = []
    for tr in table.find_all('tr'):
        cells = tr.find_all(['td', 'th'])
        if cells:
            rows.append([_clean(c.get_text()) for c in cells])
    return rows


def _read_excel_rows(file_bytes):
    """Every cell of every sheet as text. Isracard splits domestic and
    foreign-currency transactions across separate sheets."""
    if _is_xlsx(file_bytes):
        try:
            from openpyxl import load_workbook
        except ImportError:
            raise ImportError_('Excel support (openpyxl) is not installed on the server')
        wb = load_workbook(io.BytesIO(file_bytes), data_only=True, read_only=True)
        rows = []
        for name in wb.sheetnames:
            for row in wb[name].iter_rows(values_only=True):
                rows.append([_cell_to_text(c) for c in row])
        return rows

    try:
        import xlrd
    except ImportError:
        raise ImportError_('Reading .xls needs the xlrd package, which is not '
                           'installed on the server')
    book = xlrd.open_workbook(file_contents=file_bytes)
    rows = []
    for sheet in book.sheets():
        for r in range(sheet.nrows):
            row = []
            for c in range(sheet.ncols):
                cell = sheet.cell(r, c)
                if cell.ctype == xlrd.XL_CELL_DATE:
                    row.append(xlrd.xldate_as_datetime(cell.value, book.datemode)
                               .strftime('%Y-%m-%d'))
                else:
                    row.append(_cell_to_text(cell.value))
            rows.append(row)
    return rows


def _find_header_row(rows):
    """Score candidate rows by how many recognisable column names they hold.

    Taking the *first* row containing a hint is wrong: Isracard's preamble has
    lines like 'פירוט עסקאות ... לתאריך חיוב 02/08/2025', which contains תאריך
    and would win over the genuine header a few rows below.
    """
    best_idx, best_score = None, 0
    for idx, row in enumerate(rows[:40]):
        cells = [_clean(c).lower() for c in row]
        if sum(1 for c in cells if c) < 2:
            continue
        score = sum(1 for c in cells if any(hint.lower() in c for hint in HEADER_HINTS))
        if score > best_score:
            best_idx, best_score = idx, score
    if best_idx is not None:
        return best_idx
    for idx, row in enumerate(rows):
        if sum(1 for c in row if _clean(c)) >= 2:
            return idx
    return 0


def _is_noise_row(values):
    joined = ' '.join(_clean(v) for v in values).lower()
    return any(marker.lower() in joined for marker in SKIP_ROW_MARKERS)


def read_table(file_bytes, filename=None):
    """Normalise CSV / XLSX / XLS / HTML-as-XLS into (headers, rows-as-dicts)."""
    name = (filename or '').lower()

    if _is_markup(file_bytes):
        raw = _read_markup_rows(file_bytes)
    elif _is_xlsx(file_bytes) or _is_xls(file_bytes):
        raw = _read_excel_rows(file_bytes)
    elif name.endswith(('.xlsx', '.xls')):
        # Named like a spreadsheet but no recognised signature — try Excel, then
        # fall back to CSV rather than failing outright.
        try:
            raw = _read_excel_rows(file_bytes)
        except Exception:
            raw = list(csv.reader(io.StringIO(_decode_hebrew(file_bytes))))
    else:
        raw = list(csv.reader(io.StringIO(_decode_hebrew(file_bytes))))

    raw = [r for r in raw if any(_clean(c) for c in r)]   # drop blank lines
    if not raw:
        raise ImportError_('The file has no rows in it')

    start = _find_header_row(raw)
    headers = [_clean(h) for h in raw[start]]
    # De-duplicate blank/repeated header cells so dict keys stay unique
    seen = {}
    clean_headers = []
    for i, h in enumerate(headers):
        h = h or f'Column {i + 1}'
        if h in seen:
            seen[h] += 1
            h = f'{h} ({seen[h]})'
        else:
            seen[h] = 1
        clean_headers.append(h)

    if len(clean_headers) < 2:
        raise ImportError_("This doesn't look like a statement — expected a table "
                           "with separate date, description and amount columns")

    rows = []
    for row in raw[start + 1:]:
        if _is_noise_row(row):        # totals, "not yet captured" section headings
            continue
        padded = [_clean(c) for c in row] + [''] * (len(clean_headers) - len(row))
        rows.append(dict(zip(clean_headers, padded)))

    if not rows:
        raise ImportError_('Found column headers but no transactions below them')

    return clean_headers, rows


def resolve_field(row, mapped_column, kind):
    """Read a mapped column, falling back to known aliases.

    A single workbook can hold tables with different headings (Isracard keeps
    foreign-currency purchases on their own sheet under different names), so a
    column chosen from one table may be absent from another.
    """
    value = row.get(mapped_column)
    if value not in (None, ''):
        return value
    lowered = {k.lower(): v for k, v in row.items()}
    for alias in COLUMN_ALIASES.get(kind, ()):
        value = lowered.get(alias.lower())
        if value not in (None, ''):
            return value
    return ''


def sniff_columns(file_bytes, filename=None):
    """Return headers + sample rows + a best-guess column mapping, without touching the DB."""
    headers, rows = read_table(file_bytes, filename)
    if not headers:
        raise ImportError_('Empty file')
    sample = [[r.get(h, '') for h in headers] for r in rows[:5]]

    # Prefer an exact alias match, then a substring match. Aliases are ordered
    # most-specific first, so 'סכום חיוב' (the amount actually charged) beats
    # 'סכום עסקה' (the original transaction amount) on Isracard exports.
    guess = {'date': None, 'amount': None, 'merchant': None}
    lowered = [(h, h.lower()) for h in headers]
    for kind, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            for original, low in lowered:
                if low == alias.lower():
                    guess[kind] = original
                    break
            if guess[kind]:
                break
        if guess[kind]:
            continue
        for alias in aliases:
            for original, low in lowered:
                if alias.lower() in low:
                    guess[kind] = original
                    break
            if guess[kind]:
                break

    if guess['date'] is None and headers:
        guess['date'] = headers[0]
    if guess['merchant'] is None and len(headers) > 1:
        guess['merchant'] = headers[1]
    if guess['amount'] is None and len(headers) > 2:
        guess['amount'] = headers[-1]

    return {'headers': headers, 'sample_rows': sample, 'guess': guess}


def _coerce_date_text(value):
    """Excel hands back real dates; CSV hands back strings."""
    if isinstance(value, datetime):
        return value.strftime('%Y-%m-%d')
    return str(value or '').strip()


def _strip_locality(s):
    """Drop a trailing city/branch name, but never the whole merchant name."""
    changed = True
    while changed:
        changed = False
        words = s.split()
        for n in range(_MAX_LOCALITY_WORDS, 0, -1):
            if len(words) <= n:  # keep at least one word of the actual merchant
                continue
            if ' '.join(words[-n:]) in _LOCALITIES:
                s = ' '.join(words[:-n])
                changed = True
                break
    return s


def normalize_merchant(raw):
    s = (raw or '').strip().upper()
    s = _PROCESSOR_PREFIX.sub('', s)
    s = _CITY_STATE_SUFFIX.sub('', s)
    s = _STORE_NUMBER.sub(' ', s)
    s = _EMBEDDED_DATE.sub(' ', s)
    s = s.replace('*', ' ')
    s = re.sub(r'\s{2,}', ' ', s).strip()
    s = _strip_locality(s)
    return s.title() if s else 'Unknown'


def compute_dedup_hash(account_id, txn_date, signed_amount, merchant_raw):
    key = f"{account_id}|{txn_date.isoformat()}|{signed_amount:.2f}|{(merchant_raw or '').strip().lower()}"
    return hashlib.sha256(key.encode('utf-8')).hexdigest()


def _parse_date(value, known_format=None):
    value = _clean(value)
    if not value:
        return None, None
    formats = ([known_format] if known_format else []) + DATE_FORMATS
    for fmt in formats:
        if not fmt:
            continue
        try:
            return datetime.strptime(value, fmt).date(), fmt
        except ValueError:
            continue
    # Last resort: dig a date out of a longer string such as "לתאריך חיוב 02/08/2025"
    m = re.search(r'(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})', value)
    if m:
        d, mth, y = m.groups()
        for fmt in ('%d/%m/%Y', '%d/%m/%y'):
            try:
                return datetime.strptime(f'{d}/{mth}/{y}', fmt).date(), None
            except ValueError:
                continue
    return None, None


def _parse_amount(value, amount_sign):
    value = _clean(value).replace(',', '').replace('$', '').replace('₪', '')
    value = value.replace('NIS', '').replace('ש"ח', '')
    # Spaces can be thousands separators here — a non-breaking space became an
    # ordinary one during cleaning, and float() rejects either.
    value = value.replace(' ', '')
    if not value:
        return None

    negative = False
    if value.startswith('(') and value.endswith(')'):
        negative = True
        value = value[1:-1]
    if value.endswith('-'):     # RTL exports place the minus sign after the number
        negative = True
        value = value[:-1].strip()

    try:
        amount = float(value)
    except ValueError:
        return None
    if negative:
        amount = -amount
    # Normalise so positive always means "money went out".
    if amount_sign == 'negative_is_debit':
        amount = -amount
    return amount


def categorize(merchant_normalized, user_id, txn_type):
    """Match the merchant against the user's rules, restricted to matching kind."""
    rules = (CategorizationRule.query
             .filter_by(user_id=user_id)
             .join(Category, CategorizationRule.category_id == Category.id)
             .filter(Category.kind == txn_type)
             .order_by(CategorizationRule.priority.desc())
             .all())
    rules.sort(key=lambda r: (r.priority or 0, len(r.match_pattern)), reverse=True)

    haystack = merchant_normalized.upper()
    for rule in rules:
        pattern = rule.match_pattern.upper()
        matched = (haystack == pattern) if rule.match_type == 'exact' else (pattern in haystack)
        if matched:
            rule.hit_count = (rule.hit_count or 0) + 1
            return rule.category_id, 'rule'
    return None, 'default'


def parse_and_import(file_bytes, account, column_mapping, user, filename=None):
    if len(file_bytes) > MAX_FILE_BYTES:
        raise ImportError_('File too large (max 2MB)')

    _headers, rows = read_table(file_bytes, filename)
    if len(rows) > MAX_ROWS:
        raise ImportError_(f'Too many rows (max {MAX_ROWS})')

    date_col = column_mapping['date']
    amount_col = column_mapping['amount']
    merchant_col = column_mapping['merchant']
    date_format = column_mapping.get('date_format')
    amount_sign = column_mapping.get('amount_sign', 'positive_is_debit')

    seed_default_rules(user)
    uncategorized = get_uncategorized_category(user)
    default_income = get_default_income_category(user)

    imported = 0
    duplicates = 0
    min_date, max_date = None, None
    new_transactions = []
    resolved_date_format = date_format

    for row in rows:
        txn_date, used_format = _parse_date(resolve_field(row, date_col, 'date'), resolved_date_format)
        if txn_date is None:
            continue
        if resolved_date_format is None:
            resolved_date_format = used_format

        signed = _parse_amount(resolve_field(row, amount_col, 'amount'), amount_sign)
        if signed is None or signed == 0:
            continue

        raw_merchant = resolve_field(row, merchant_col, 'merchant') or ''
        merchant_normalized = normalize_merchant(raw_merchant)
        dedup_hash = compute_dedup_hash(account.id, txn_date, signed, raw_merchant)

        if Transaction.query.filter_by(account_id=account.id, dedup_hash=dedup_hash).first():
            duplicates += 1
            continue

        # Money coming in shows up as a credit on the statement.
        txn_type = 'income' if signed < 0 else 'expense'
        category_id, source = categorize(merchant_normalized, user.id, txn_type)
        if category_id is None:
            category_id = default_income.id if txn_type == 'income' else uncategorized.id

        new_transactions.append(Transaction(
            user_id=user.id,
            account_id=account.id,
            category_id=category_id,
            date=txn_date,
            amount=abs(signed),
            txn_type=txn_type,
            merchant_raw=raw_merchant.strip(),
            merchant_normalized=merchant_normalized,
            category_source=source,
            dedup_hash=dedup_hash,
        ))
        imported += 1

        if min_date is None or txn_date < min_date:
            min_date = txn_date
        if max_date is None or txn_date > max_date:
            max_date = txn_date

    import_record = Import(
        user_id=user.id, account_id=account.id, filename=filename,
        row_count=len(rows), imported_count=imported, duplicate_count=duplicates,
        date_range_start=min_date, date_range_end=max_date,
    )
    db.session.add(import_record)
    db.session.flush()

    for txn in new_transactions:
        txn.import_id = import_record.id
        db.session.add(txn)

    saved_mapping = dict(column_mapping)
    saved_mapping['date_format'] = resolved_date_format
    account.column_mapping = json.dumps(saved_mapping)

    db.session.commit()
    return import_record


def apply_correction(transaction, new_category_id, user, retroactive=True):
    """Recategorise a transaction and remember the choice for future imports."""
    category = Category.query.filter_by(id=new_category_id, user_id=user.id).first()
    if not category:
        raise ImportError_('Unknown category')

    transaction.category_id = category.id
    transaction.category_source = 'manual'
    transaction.txn_type = category.kind

    pattern = transaction.merchant_normalized
    rule = CategorizationRule.query.filter_by(user_id=user.id, match_pattern=pattern).first()
    if rule:
        rule.category_id = category.id
        rule.priority = 10
    else:
        db.session.add(CategorizationRule(
            user_id=user.id, category_id=category.id,
            match_pattern=pattern, match_type='contains', priority=10,
        ))

    if retroactive:
        (Transaction.query
         .filter(Transaction.user_id == user.id,
                 Transaction.merchant_normalized == pattern,
                 Transaction.category_source != 'manual',
                 Transaction.id != transaction.id)
         .update({'category_id': category.id, 'category_source': 'rule',
                  'txn_type': category.kind}, synchronize_session=False))

    db.session.commit()
